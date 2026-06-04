import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font

# =========================================================
# PATHS
# =========================================================

ROOT = Path(__file__).resolve().parent.parent

DATA_DIR = ROOT / "data"

OUTPUT_FILE = (
    DATA_DIR
    / "monthly_factsheet.xlsx"
)

# =========================================================
# HELPER
# =========================================================

def safe_load(file_name):

    file_path = DATA_DIR / file_name

    if not file_path.exists():

        print(
            f"⚠ Missing: {file_name}"
        )

        return None

    try:

        return pd.read_csv(
            file_path
        )

    except Exception as e:

        print(
            f"⚠ Failed loading {file_name}: {e}"
        )

        return None

# =========================================================
# LOAD FILES
# =========================================================

print(
    "\n📥 Loading Factsheet Inputs..."
)

regime_df = safe_load(
    "market_regime.csv"
)

portfolio_df = safe_load(
    "optimised_portfolio.csv"
)

risk_df = safe_load(
    "portfolio_risk_report.csv"
)

factor_df = safe_load(
    "performance_attribution.csv"
)

rebalance_df = safe_load(
    "rebalance_plan.csv"
)

stoploss_df = safe_load(
    "stoploss_signals.csv"
)

stats_df = safe_load(
    "walk_forward_stats.csv"
)

# =========================================================
# WORKBOOK
# =========================================================

wb = Workbook()

cover = wb.active

cover.title = "Factsheet"

# =========================================================
# COVER PAGE
# =========================================================

cover["A1"] = (
    "Institutional Quant Alpha"
)

cover["A1"].font = Font(
    size=18,
    bold=True
)

cover["A3"] = (
    f"Generated: "
    f"{datetime.now():%Y-%m-%d %H:%M}"
)

regime = "UNKNOWN"

if (
    regime_df is not None
    and "REGIME" in regime_df.columns
    and len(regime_df) > 0
):

    regime = str(
        regime_df["REGIME"].iloc[0]
    )

cover["A5"] = (
    f"Current Regime: {regime}"
)

cover["A7"] = (
    f"Portfolio Holdings: "
    f"{len(portfolio_df) if portfolio_df is not None else 0}"
)

cover["A8"] = (
    f"Risk Records: "
    f"{len(risk_df) if risk_df is not None else 0}"
)

cover["A9"] = (
    f"Rebalance Actions: "
    f"{len(rebalance_df) if rebalance_df is not None else 0}"
)

# =========================================================
# PERFORMANCE
# =========================================================

perf_sheet = wb.create_sheet(
    "Performance"
)

perf_sheet["A1"] = (
    "Performance Metrics"
)

perf_sheet["A1"].font = Font(
    bold=True
)

if stats_df is not None:

    for c, col in enumerate(
        stats_df.columns,
        start=1
    ):

        perf_sheet.cell(
            2,
            c,
            col
        )

    for r, row in enumerate(
        stats_df.values,
        start=3
    ):

        for c, value in enumerate(
            row,
            start=1
        ):

            perf_sheet.cell(
                r,
                c,
                value
            )

# =========================================================
# PORTFOLIO
# =========================================================

portfolio_sheet = wb.create_sheet(
    "Portfolio"
)

portfolio_sheet["A1"] = (
    "Top Holdings"
)

portfolio_sheet["A1"].font = Font(
    bold=True
)

if portfolio_df is not None:

    weight_col = None

    if "OPTIMAL_WEIGHT" in portfolio_df.columns:

        weight_col = "OPTIMAL_WEIGHT"

    elif "FINAL_WEIGHT" in portfolio_df.columns:

        weight_col = "FINAL_WEIGHT"

    cols = [

        c

        for c in [

            "Symbol",
            weight_col,
            "EXPECTED_RETURN_30D",
            "CONVICTION_SCORE"

        ]

        if c is not None
        and c in portfolio_df.columns

    ]

    if cols:

        subset = portfolio_df[
            cols
        ].head(20)

        for col_num, col in enumerate(
            subset.columns,
            start=1
        ):

            portfolio_sheet.cell(
                3,
                col_num,
                col
            )

        for row_num, row in enumerate(
            subset.values,
            start=4
        ):

            for col_num, value in enumerate(
                row,
                start=1
            ):

                portfolio_sheet.cell(
                    row_num,
                    col_num,
                    value
                )

# =========================================================
# RISK
# =========================================================

risk_sheet = wb.create_sheet(
    "Risk"
)

risk_sheet["A1"] = (
    "Portfolio Risk"
)

risk_sheet["A1"].font = Font(
    bold=True
)

if risk_df is not None:

    for c, col in enumerate(
        risk_df.columns,
        start=1
    ):

        risk_sheet.cell(
            2,
            c,
            col
        )

    for r, row in enumerate(
        risk_df.values,
        start=3
    ):

        for c, value in enumerate(
            row,
            start=1
        ):

            risk_sheet.cell(
                r,
                c,
                value
            )

# =========================================================
# FACTOR ATTRIBUTION
# =========================================================

factor_sheet = wb.create_sheet(
    "Attribution"
)

factor_sheet["A1"] = (
    "Factor Attribution"
)

factor_sheet["A1"].font = Font(
    bold=True
)

if factor_df is not None:

    for c, col in enumerate(
        factor_df.columns,
        start=1
    ):

        factor_sheet.cell(
            2,
            c,
            col
        )

    for r, row in enumerate(
        factor_df.values,
        start=3
    ):

        for c, value in enumerate(
            row,
            start=1
        ):

            factor_sheet.cell(
                r,
                c,
                value
            )

# =========================================================
# REBALANCE
# =========================================================

rebalance_sheet = wb.create_sheet(
    "Rebalance"
)

rebalance_sheet["A1"] = (
    "Rebalance Actions"
)

rebalance_sheet["A1"].font = Font(
    bold=True
)

rebalance_sheet.freeze_panes = "A4"

if rebalance_df is not None:

    for c, col in enumerate(
        rebalance_df.columns,
        start=1
    ):

        rebalance_sheet.cell(
            3,
            c,
            col
        )

    for r, row in enumerate(
        rebalance_df.values,
        start=4
    ):

        for c, value in enumerate(
            row,
            start=1
        ):

            rebalance_sheet.cell(
                r,
                c,
                value
            )

# =========================================================
# STOP LOSS
# =========================================================

stop_sheet = wb.create_sheet(
    "StopLoss"
)

stop_sheet["A1"] = (
    "Stop Loss Signals"
)

stop_sheet["A1"].font = Font(
    bold=True
)

stop_sheet.freeze_panes = "A4"

if stoploss_df is not None:

    for c, col in enumerate(
        stoploss_df.columns,
        start=1
    ):

        stop_sheet.cell(
            3,
            c,
            col
        )

    for r, row in enumerate(
        stoploss_df.values,
        start=4
    ):

        for c, value in enumerate(
            row,
            start=1
        ):

            stop_sheet.cell(
                r,
                c,
                value
            )

# =========================================================
# AUTO WIDTH
# =========================================================

for sheet in wb.worksheets:

    for col in sheet.columns:

        width = max(

            len(str(cell.value))
            if cell.value is not None
            else 0

            for cell in col

        )

        sheet.column_dimensions[
            col[0].column_letter
        ].width = min(
            max(width + 4, 12),
            60
        )

# =========================================================
# SAVE
# =========================================================

wb.save(
    OUTPUT_FILE
)

print(
    "\n✅ Monthly Factsheet Generated"
)

print(
    "\n📁 Saved:"
)

print(
    OUTPUT_FILE
)